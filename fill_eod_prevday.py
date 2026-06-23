#!/usr/bin/env python3
"""
fill_eod_prevday.py
-------------------
Pull the PREVIOUS trading day's OHLCV for every ticker/date in the four EOD
Winners/Losers tabs and write prev-day signal columns into the tracker workbook.

Runs on Render (where data.alpaca.markets is reachable). Reads Alpaca keys from
environment variables — never hardcode them:

    export ALPACA_API_KEY_ID=...        # your Alpaca key id
    export ALPACA_API_SECRET_KEY=...    # your Alpaca secret

(Reminder: revoke any previously-exposed tokens, and do NOT commit keys.)

New columns appended to the right of each EOD tab (matching existing header style):
    Prev High         prior trading day's high
    Prev Low          prior trading day's low
    Prev Volume       prior trading day's volume
    Prev Range %      (prev_high - prev_low) / prev_close * 100
    Prev HOD>Close %  (prev_high - prev_close) / prev_close * 100   <-- EXHAUSTION SIGNAL (>25 = exhausted)
    Prev Close Pos    (prev_close - prev_low) / (prev_high - prev_low)   1=closed strong, 0=closed weak
    Prev Data OK      OK | GLITCH | NO_DATA | CHECK(close mismatch vs sheet Prev Close)

Existing data is never overwritten. Invalid/missing prev-day data leaves the
numeric cells blank (only the status column is set), per the tracker's blank-if-
not-provided rule. Glitch filter mirrors the EOD rule: skip bars where
prev_close > 2x prev_high or OHLC is impossible (bad IEX feed prints).

Usage:
    python fill_eod_prevday.py --in merged_trading_tracker__22-06-26_.xlsx
    python fill_eod_prevday.py --in <file> --out <file> --feed iex
    python fill_eod_prevday.py --in <file> --dry-run            # no Alpaca calls; validate + preview headers
    python fill_eod_prevday.py --in <file> --limit-tickers 50   # small test run
"""
import os, sys, time, argparse, datetime as dt, copy
from collections import defaultdict
import openpyxl

ALPACA_BARS_URL = "https://data.alpaca.markets/v2/stocks/bars"
EOD_TABS = ['EOD Biggest Winners Q1', 'EOD Biggest Losers Q1',
            'EOD Biggest Winners Q2', 'EOD Biggest Losers Q2']
NEW_COLS = ['Prev High', 'Prev Low', 'Prev Volume', 'Prev Range %',
            'Prev HOD>Close %', 'Prev Close Pos', 'Prev Data OK']

# ----------------------------- workbook helpers -----------------------------

def find_header_row(ws):
    """Header row = the row whose col A == 'Date' and col B == 'Ticker'.
    Handles the Losers-Q2 one-row offset automatically."""
    for r in range(1, 9):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if str(a).strip() == 'Date' and str(b).strip() == 'Ticker':
            return r
    raise ValueError(f"Could not find Date/Ticker header in sheet {ws.title!r}")

def _to_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y"):
            try:
                return dt.datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None

def collect_pairs(wb):
    """Walk every EOD tab and return (tab, row, ticker, date) records."""
    pairs, tickers = [], set()
    dmin = dmax = None
    for tab in EOD_TABS:
        ws = wb[tab]
        hdr = find_header_row(ws)
        r = hdr + 1
        blanks = 0
        while blanks < 3:                     # stop after 3 consecutive blank tickers
            tk = ws.cell(r, 2).value
            if tk in (None, ''):
                blanks += 1; r += 1; continue
            blanks = 0
            d = _to_date(ws.cell(r, 1).value)
            if d:
                t = str(tk).strip().upper()
                pairs.append((tab, r, t, d))
                tickers.add(t)
                dmin = d if (dmin is None or d < dmin) else dmin
                dmax = d if (dmax is None or d > dmax) else dmax
            r += 1
    return pairs, tickers, dmin, dmax

def last_used_col(ws, row):
    last = 1
    for c in range(1, 80):
        if ws.cell(row, c).value not in (None, ''):
            last = c
    return last

def ensure_headers(ws, hdr):
    """Find existing prev-day columns or create them to the right.
    Returns the starting column index for NEW_COLS."""
    last = last_used_col(ws, hdr)
    existing = {ws.cell(hdr, c).value: c for c in range(1, last + 1)}
    if NEW_COLS[0] in existing:
        return existing[NEW_COLS[0]]
    start = last + 1
    src = ws.cell(hdr, last)                    # copy look from an existing header cell
    for j, name in enumerate(NEW_COLS):
        c = ws.cell(hdr, start + j)
        c.value = name
        c.font = copy.copy(src.font)
        c.fill = copy.copy(src.fill)
        c.alignment = copy.copy(src.alignment)
        c.border = copy.copy(src.border)
    return start

# ------------------------------- alpaca fetch -------------------------------

def _get_with_retry(url, params, hdr, tries=5):
    import requests                             # lazy import so --dry-run needs no network libs
    for t in range(tries):
        r = requests.get(url, params=params, headers=hdr, timeout=30)
        if r.status_code == 200:
            return r
        if r.status_code == 429:
            wait = 2 ** t
            print(f"  rate-limited, sleeping {wait}s", flush=True); time.sleep(wait); continue
        if r.status_code in (500, 502, 503, 504):
            time.sleep(2 ** t); continue
        raise RuntimeError(f"Alpaca {r.status_code}: {r.text[:300]}")
    raise RuntimeError("Alpaca: max retries exceeded")

def fetch_bars(symbols, start, end, key, secret, feed='iex', chunk=100, pause=0.3):
    """Multi-symbol daily bars -> {symbol: {date: bar}}. Chunks symbols and paginates."""
    hdr = {'APCA-API-KEY-ID': key, 'APCA-API-SECRET-KEY': secret}
    out = defaultdict(dict)
    syms = sorted(symbols)
    for i in range(0, len(syms), chunk):
        batch = syms[i:i + chunk]
        token = None
        while True:
            params = {'symbols': ','.join(batch), 'timeframe': '1Day',
                      'start': start.isoformat(), 'end': end.isoformat(),
                      'feed': feed, 'adjustment': 'raw', 'limit': 10000}
            if token:
                params['page_token'] = token
            js = _get_with_retry(ALPACA_BARS_URL, params, hdr).json()
            for sym, bars in (js.get('bars') or {}).items():
                for b in bars:
                    d = dt.datetime.fromisoformat(b['t'].replace('Z', '+00:00')).date()
                    out[sym][d] = b
            token = js.get('next_page_token')
            if not token:
                break
            time.sleep(pause)
        print(f"  fetched {min(i + chunk, len(syms))}/{len(syms)} symbols", flush=True)
        time.sleep(pause)
    return out

def prev_trading_bar(bars_by_date, target):
    """Latest bar strictly before target date (Alpaca returns only trading days)."""
    prior = [d for d in bars_by_date if d < target]
    return bars_by_date[max(prior)] if prior else None

# ----------------------------------- main -----------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--in', dest='inp', default='merged_trading_tracker__22-06-26_.xlsx')
    ap.add_argument('--out', dest='out', default=None)
    ap.add_argument('--feed', default='iex', help="iex (free) or sip (paid)")
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--preview', action='store_true', help="with --dry-run, save a headers-only preview workbook")
    ap.add_argument('--limit-tickers', type=int, default=0)
    a = ap.parse_args()
    out = a.out or a.inp.replace('.xlsx', '') + f"__prevday_{dt.date.today():%d-%m-%y}.xlsx"

    print("Loading workbook…", flush=True)
    wb = openpyxl.load_workbook(a.inp)
    pairs, tickers, dmin, dmax = collect_pairs(wb)
    print(f"Rows: {len(pairs)} | unique tickers: {len(tickers)} | date span: {dmin} .. {dmax}", flush=True)

    if a.limit_tickers:
        keep = set(sorted(tickers)[:a.limit_tickers])
        pairs = [p for p in pairs if p[2] in keep]
        tickers = keep
        print(f"  limited to {len(tickers)} tickers / {len(pairs)} rows", flush=True)

    print(f"Estimated API chunks (~100 symbols each): {(len(tickers) + 99) // 100}", flush=True)

    # build header caches (also creates the new columns)
    hdr_cache, start_cache = {}, {}
    for tab in EOD_TABS:
        ws = wb[tab]
        hdr_cache[tab] = find_header_row(ws)
        start_cache[tab] = ensure_headers(ws, hdr_cache[tab])
        print(f"  {tab}: header row {hdr_cache[tab]}, new cols start at col {start_cache[tab]}", flush=True)

    if a.dry_run:
        if a.preview:
            wb.save(out)
            print(f"DRY RUN — saved headers-only preview to {out}")
        else:
            print("DRY RUN — parsing/headers validated, no Alpaca calls, nothing saved.")
        return

    key = os.environ.get('ALPACA_API_KEY_ID')
    secret = os.environ.get('ALPACA_API_SECRET_KEY')
    if not key or not secret:
        sys.exit("ERROR: set ALPACA_API_KEY_ID and ALPACA_API_SECRET_KEY env vars (do not hardcode).")

    start = dmin - dt.timedelta(days=14)        # buffer so the prior trading day is always covered
    print("Fetching daily bars from Alpaca…", flush=True)
    bars = fetch_bars(tickers, start, dmax, key, secret, feed=a.feed)

    filled = glitch = missing = mismatch = 0
    for (tab, r, tk, d) in pairs:
        ws = wb[tab]
        sc = start_cache[tab]
        # number formats sampled from existing cells in this row
        fmt_price = ws.cell(r, 20).number_format or '0.0000'   # High col
        fmt_vol = ws.cell(r, 9).number_format or '#,##0'       # Volume col

        pb = prev_trading_bar(bars.get(tk, {}), d)
        vals = [None] * 7
        if pb is None:
            vals[6] = 'NO_DATA'; missing += 1
        else:
            ph, pl, pc, pv = pb['h'], pb['l'], pb['c'], pb['v']
            if not (ph >= pl > 0 and pc > 0) or pc > 2 * ph:
                vals[6] = 'GLITCH'; glitch += 1
            else:
                rng = (ph - pl) / pc * 100
                hodac = (ph - pc) / pc * 100
                cpos = (pc - pl) / (ph - pl) if ph > pl else None
                vals = [round(ph, 4), round(pl, 4), int(pv), round(rng, 2),
                        round(hodac, 2), round(cpos, 3) if cpos is not None else None, 'OK']
                filled += 1
                exist_pc = ws.cell(r, 18).value                # cross-check sheet's Prev Close
                if isinstance(exist_pc, (int, float)) and exist_pc > 0 and abs(exist_pc - pc) / pc > 0.05:
                    vals[6] = 'CHECK'; mismatch += 1

        for j, v in enumerate(vals):
            cell = ws.cell(r, sc + j)
            if v is None:
                continue
            cell.value = v
            if j in (0, 1):
                cell.number_format = fmt_price
            elif j == 2:
                cell.number_format = fmt_vol
            elif j in (3, 4):
                cell.number_format = '0.00'
            elif j == 5:
                cell.number_format = '0.000'

    print(f"\nDone. filled={filled}  glitch={glitch}  no_data={missing}  close_mismatch(CHECK)={mismatch}", flush=True)
    wb.save(out)
    print(f"Saved -> {out}", flush=True)


if __name__ == '__main__':
    main()
