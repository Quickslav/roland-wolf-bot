"""
fill_eod_premarket.py

Fetches pre-market data (4:00-9:29 AM ET) for every ticker in the
EOD Biggest Winners and EOD Biggest Losers tabs of the trading tracker.

Adds 4 new columns at the end of each tab:
  PM Open, PM High, PM Low, PM Volume

Only fills cells that are blank. Re-running is safe.

Usage:
  python fill_eod_premarket.py --file "merged_trading_tracker_eod.xlsx"

Optional flags:
  --winners-only      Only process EOD Biggest Winners
  --losers-only       Only process EOD Biggest Losers
  --start YYYY-MM-DD  Only process rows on or after this date
  --end YYYY-MM-DD    Only process rows on or before this date
"""

import argparse
import time
from datetime import datetime, timedelta, timezone

import openpyxl
import requests

# --- Alpaca credentials ---
API_KEY    = "PKIYGXQT3DGX7B6BDIZFZ6VQWU"
API_SECRET = "Ekaz5bQHUbbFUQvmidbBSU89wHMtgigik3TsyFD15NA3"
DATA_URL   = "https://data.alpaca.markets/v2"

HEADERS = {
    "APCA-API-KEY-ID":     API_KEY,
    "APCA-API-SECRET-KEY": API_SECRET,
}

# --- Tab layout ---
# Header row = where the column titles live; data_start = first ticker row.
TAB_LAYOUTS = {
    "EOD Biggest Winners Q1": {"header_row": 3, "data_start": 4},
    "EOD Biggest Losers Q1":  {"header_row": 3, "data_start": 4},
    "EOD Biggest Winners Q2": {"header_row": 3, "data_start": 4},
    "EOD Biggest Losers Q2":  {"header_row": 4, "data_start": 5},
}

# New columns appended at the end (after existing 23)
PM_OPEN_COL   = 24
PM_HIGH_COL   = 25
PM_LOW_COL    = 26
PM_VOLUME_COL = 27

PM_HEADERS = {
    PM_OPEN_COL:   "PM Open",
    PM_HIGH_COL:   "PM High",
    PM_LOW_COL:    "PM Low",
    PM_VOLUME_COL: "PM Volume",
}


def fetch_premarket_bars(symbol, date):
    """Fetch 1-minute bars between 4:00 AM and 9:29 AM ET for a date."""
    # ET is UTC-4 (EDT) for most of the year; use ISO with -04:00
    start = f"{date.strftime('%Y-%m-%d')}T04:00:00-04:00"
    end   = f"{date.strftime('%Y-%m-%d')}T09:29:00-04:00"
    url = f"{DATA_URL}/stocks/{symbol}/bars"
    params = {
        "timeframe": "1Min",
        "start":     start,
        "end":       end,
        "limit":     1000,
        "feed":      "sip",  # SIP feed required for pre-market
        "adjustment": "raw",
    }
    try:
        r = requests.get(url, headers=HEADERS, params=params, timeout=10)
        if r.status_code != 200:
            return None
        bars = r.json().get("bars", [])
        return bars if bars else None
    except Exception:
        return None


def aggregate_pm(bars):
    """From minute bars, derive PM open, high, low, volume."""
    if not bars:
        return None, None, None, None
    pm_open = bars[0]["o"]
    pm_high = max(b["h"] for b in bars)
    pm_low  = min(b["l"] for b in bars)
    pm_vol  = sum(b["v"] for b in bars)
    return pm_open, pm_high, pm_low, pm_vol


def process_tab(ws, layout, start_date=None, end_date=None):
    """Process one EOD tab. Returns (rows_processed, cells_filled)."""
    header_row = layout["header_row"]
    data_start = layout["data_start"]

    # Ensure header cells exist
    for col, label in PM_HEADERS.items():
        cell = ws.cell(header_row, col)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = label

    # Auto-widen new columns
    ws.column_dimensions[openpyxl.utils.get_column_letter(PM_OPEN_COL)].width   = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(PM_HIGH_COL)].width   = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(PM_LOW_COL)].width    = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(PM_VOLUME_COL)].width = 14

    rows_processed = 0
    cells_filled   = 0

    for row_idx in range(data_start, ws.max_row + 1):
        date_val = ws.cell(row_idx, 1).value
        ticker   = ws.cell(row_idx, 2).value

        if not date_val or not ticker:
            continue
        if isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val[:10], "%Y-%m-%d")
            except ValueError:
                continue
        # Filter by date range if provided
        if start_date and date_val.date() < start_date:
            continue
        if end_date and date_val.date() > end_date:
            continue

        # Check if any PM cells are blank
        cells_to_fill = [c for c in PM_HEADERS if ws.cell(row_idx, c).value in (None, "")]
        if not cells_to_fill:
            continue  # already filled

        bars = fetch_premarket_bars(ticker.strip().upper(), date_val)
        pm_open, pm_high, pm_low, pm_vol = aggregate_pm(bars)

        if PM_OPEN_COL in cells_to_fill and pm_open is not None:
            ws.cell(row_idx, PM_OPEN_COL).value = pm_open
            ws.cell(row_idx, PM_OPEN_COL).number_format = "#,##0.0000"
            cells_filled += 1
        if PM_HIGH_COL in cells_to_fill and pm_high is not None:
            ws.cell(row_idx, PM_HIGH_COL).value = pm_high
            ws.cell(row_idx, PM_HIGH_COL).number_format = "#,##0.0000"
            cells_filled += 1
        if PM_LOW_COL in cells_to_fill and pm_low is not None:
            ws.cell(row_idx, PM_LOW_COL).value = pm_low
            ws.cell(row_idx, PM_LOW_COL).number_format = "#,##0.0000"
            cells_filled += 1
        if PM_VOLUME_COL in cells_to_fill and pm_vol is not None:
            ws.cell(row_idx, PM_VOLUME_COL).value = pm_vol
            ws.cell(row_idx, PM_VOLUME_COL).number_format = "#,##0"
            cells_filled += 1

        rows_processed += 1
        if rows_processed % 50 == 0:
            print(f"  ...{rows_processed} rows processed, {cells_filled} cells filled")
        time.sleep(0.3)  # rate limit

    return rows_processed, cells_filled


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="Excel tracker filename")
    ap.add_argument("--tab", help="Process only this tab name (e.g. 'EOD Biggest Winners Q1')")
    ap.add_argument("--start", help="YYYY-MM-DD")
    ap.add_argument("--end",   help="YYYY-MM-DD")
    args = ap.parse_args()

    start_date = datetime.strptime(args.start, "%Y-%m-%d").date() if args.start else None
    end_date   = datetime.strptime(args.end,   "%Y-%m-%d").date() if args.end   else None

    print(f"Loading {args.file}...")
    wb = openpyxl.load_workbook(args.file)

    tabs_to_run = list(TAB_LAYOUTS.keys())
    if args.tab:
        if args.tab not in TAB_LAYOUTS:
            print(f"Unknown tab '{args.tab}'. Valid: {list(TAB_LAYOUTS.keys())}")
            return
        tabs_to_run = [args.tab]

    for tab_name in tabs_to_run:
        if tab_name not in wb.sheetnames:
            print(f"  [skip] {tab_name} not in workbook")
            continue
        print(f"\nProcessing {tab_name}...")
        ws = wb[tab_name]
        rows, cells = process_tab(ws, TAB_LAYOUTS[tab_name], start_date, end_date)
        print(f"  Done: {rows} rows touched, {cells} cells filled.")
        # Save incrementally so a crash doesn't lose progress
        wb.save(args.file)
        print(f"  Saved {args.file}.")

    print("\nAll done.")


if __name__ == "__main__":
    main()
