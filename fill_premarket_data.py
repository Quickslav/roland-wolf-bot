"""
fill_premarket_data.py

Fills the following columns in Pre-Trade Data sheet from Alpaca:
  M  (col 13) - Prev High ($)
  N  (col 14) - Prev Low ($)
  P  (col 16) - PM Open (pre-market open, first bar 4:00-9:30 AM ET)
  S  (col 19) - Market Open (9:30 AM ET open price)
  T  (col 20) - Day Range (HOD - LOD, calculated)
  U  (col 21) - High of Day ($)
  V  (col 22) - Low of Day ($)
  W  (col 23) - Close Price ($)

Only fills rows where the cell is currently blank.
Skips rows with no ticker or no date.

Usage:
    python fill_premarket_data.py --file merged_trading_tracker.xlsx
    python fill_premarket_data.py --file merged_trading_tracker.xlsx --start-row 295
    python fill_premarket_data.py --file merged_trading_tracker.xlsx --dry-run

Requires:
    pip install openpyxl requests pytz
"""

import argparse
import time
import requests
import pytz
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Alpaca credentials ──────────────────────────────────────────────────────
ALPACA_API_KEY    = "PKRPWB8R7ATV1FFPPF7B"
ALPACA_SECRET_KEY = "FoIEtLBVCLbEjvCPRLjlBt8jbf3MWfuAzO5bCVL4"
ALPACA_DATA_URL   = "https://data.alpaca.markets"

HEADERS = {
    "APCA-API-KEY-ID":     ALPACA_API_KEY,
    "APCA-API-SECRET-KEY": ALPACA_SECRET_KEY,
}

ET = pytz.timezone("America/New_York")

# ── Column indices (1-based) ─────────────────────────────────────────────────
COL_DATE       = 1
COL_TICKER     = 2
COL_PREV_HIGH  = 13   # M
COL_PREV_LOW   = 14   # N
COL_PM_OPEN    = 16   # P
COL_MKT_OPEN   = 19   # S
COL_DAY_RANGE  = 20   # T
COL_HOD        = 21   # U
COL_LOD        = 22   # V
COL_CLOSE      = 23   # W

SHEET_NAME = "Pre-Trade Data"
HEADER_ROW = 2   # row 2 is headers, data starts row 3


def get_daily_bar(ticker: str, trade_date: date):
    """
    Fetch the daily bar for ticker on trade_date and the previous trading day.
    Returns (trade_bar, prev_bar) as dicts or (None, None) on failure.
    """
    # Request a window of 5 calendar days before and including trade_date
    # to safely capture the previous trading day
    start = (trade_date - timedelta(days=7)).isoformat()
    end   = trade_date.isoformat()

    url = f"{ALPACA_DATA_URL}/v2/stocks/{ticker}/bars"
    params = {
        "timeframe": "1Day",
        "start":     start,
        "end":       end,
        "feed":      "sip",
        "limit":     10,
        "adjustment": "raw",
    }
    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=10)
        if resp.status_code == 422:
            # Crypto or unsupported ticker
            return None, None
        resp.raise_for_status()
        bars = resp.json().get("bars", [])
        if not bars:
            return None, None

        # bars are sorted oldest-first
        # Find bar matching trade_date
        trade_bar = None
        prev_bar  = None
        for i, b in enumerate(bars):
            bar_date = datetime.fromisoformat(b["t"].replace("Z","")).date()
            if bar_date == trade_date:
                trade_bar = b
                if i > 0:
                    prev_bar = bars[i - 1]
                break

        return trade_bar, prev_bar

    except Exception as e:
        print(f"    [daily bar error] {ticker} {trade_date}: {e}")
        return None, None


def get_intraday_bars(ticker: str, trade_date: date):
    """
    Fetch 1-minute intraday bars for ticker on trade_date (ET timezone).
    Returns list of bar dicts or [] on failure.
    """
    # Build ET-aware datetime window: 4:00 AM to 8:00 PM ET
    start_et = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 4, 0, 0))
    end_et   = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 20, 0, 0))

    url = f"{ALPACA_DATA_URL}/v2/stocks/{ticker}/bars"
    params = {
        "timeframe": "1Min",
        "start":     start_et.isoformat(),
        "end":       end_et.isoformat(),
        "feed":      "sip",
        "limit":     1000,
        "adjustment": "raw",
    }
    bars = []
    try:
        while True:
            resp = requests.get(url, headers=HEADERS, params=params, timeout=15)
            if resp.status_code == 422:
                return []
            resp.raise_for_status()
            data = resp.json()
            bars.extend(data.get("bars", []))
            next_token = data.get("next_page_token")
            if not next_token:
                break
            params["page_token"] = next_token
    except Exception as e:
        print(f"    [intraday error] {ticker} {trade_date}: {e}")

    return bars


def extract_pm_open(bars: list, trade_date: date):
    """
    Pre-market open = open price of the first 1-min bar between
    4:00 AM and 9:29 AM ET on trade_date.
    """
    mkt_open_et = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 9, 30, 0))
    pm_start_et = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 4, 0, 0))

    for b in bars:
        # Parse timestamp - Alpaca returns UTC
        ts = datetime.fromisoformat(b["t"].replace("Z", "+00:00"))
        ts_et = ts.astimezone(ET)
        if pm_start_et <= ts_et < mkt_open_et:
            return b["o"]
    return None


def extract_market_open(bars: list, trade_date: date):
    """
    Market open = open price of the first 1-min bar at or after 9:30 AM ET.
    """
    mkt_open_et = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 9, 30, 0))
    mkt_close_et = ET.localize(datetime(trade_date.year, trade_date.month, trade_date.day, 16, 0, 0))

    for b in bars:
        ts = datetime.fromisoformat(b["t"].replace("Z", "+00:00"))
        ts_et = ts.astimezone(ET)
        if mkt_open_et <= ts_et < mkt_close_et:
            return b["o"]
    return None


def needs_filling(ws, row: int) -> bool:
    """Returns True if any of the target columns are blank."""
    targets = [COL_PREV_HIGH, COL_PREV_LOW, COL_PM_OPEN,
               COL_MKT_OPEN, COL_DAY_RANGE, COL_HOD, COL_LOD, COL_CLOSE]
    return any(ws.cell(row=row, column=c).value is None for c in targets)


def main():
    parser = argparse.ArgumentParser(description="Fill Pre-Trade Data columns from Alpaca")
    parser.add_argument("--file",      required=True, help="Path to the .xlsx file")
    parser.add_argument("--start-row", type=int, default=3, help="First data row to process (default: 3)")
    parser.add_argument("--end-row",   type=int, default=None, help="Last data row to process (default: last row with data)")
    parser.add_argument("--dry-run",   action="store_true", help="Print what would be filled without saving")
    parser.add_argument("--delay",     type=float, default=0.3, help="Seconds between Alpaca requests (default: 0.3)")
    args = parser.parse_args()

    print(f"Loading {args.file} ...")
    wb = load_workbook(args.file)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found.")
        return

    ws = wb[SHEET_NAME]

    # Determine end row
    last_row = args.start_row
    for row in ws.iter_rows(min_row=args.start_row, max_row=ws.max_row):
        if any(cell.value is not None for cell in row):
            last_row = row[0].row
    end_row = args.end_row or last_row

    print(f"Processing rows {args.start_row} – {end_row} ...")
    print()

    filled_count = 0
    skipped_count = 0
    error_count = 0

    # Cache intraday bars per (ticker, date) to avoid double-fetching
    intraday_cache = {}

    for row_num in range(args.start_row, end_row + 1):
        ticker = ws.cell(row=row_num, column=COL_TICKER).value
        raw_date = ws.cell(row=row_num, column=COL_DATE).value

        if not ticker or not raw_date:
            continue

        # Parse date
        if isinstance(raw_date, datetime):
            trade_date = raw_date.date()
        elif isinstance(raw_date, date):
            trade_date = raw_date
        else:
            try:
                trade_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
            except Exception:
                print(f"  Row {row_num}: Cannot parse date '{raw_date}', skipping.")
                skipped_count += 1
                continue

        # Skip weekends and known holidays
        if trade_date.weekday() >= 5:
            skipped_count += 1
            continue

        if not needs_filling(ws, row_num):
            skipped_count += 1
            continue

        print(f"  Row {row_num}: {trade_date} {ticker}")

        # ── Daily bars ────────────────────────────────────────────────────────
        trade_bar, prev_bar = get_daily_bar(ticker, trade_date)
        time.sleep(args.delay)

        # ── Intraday bars ─────────────────────────────────────────────────────
        cache_key = (ticker, trade_date)
        if cache_key not in intraday_cache:
            intraday_bars = get_intraday_bars(ticker, trade_date)
            intraday_cache[cache_key] = intraday_bars
            time.sleep(args.delay)
        else:
            intraday_bars = intraday_cache[cache_key]

        # ── Extract values ────────────────────────────────────────────────────
        prev_high  = round(prev_bar["h"],  4) if prev_bar else None
        prev_low   = round(prev_bar["l"],  4) if prev_bar else None
        hod        = round(trade_bar["h"], 4) if trade_bar else None
        lod        = round(trade_bar["l"], 4) if trade_bar else None
        close      = round(trade_bar["c"], 4) if trade_bar else None
        day_range  = round(hod - lod, 4)      if (hod is not None and lod is not None) else None
        pm_open    = round(extract_pm_open(intraday_bars, trade_date),     4) if intraday_bars else None
        mkt_open   = round(extract_market_open(intraday_bars, trade_date), 4) if intraday_bars else None

        # ── Log ───────────────────────────────────────────────────────────────
        print(f"    prev_high={prev_high}  prev_low={prev_low}")
        print(f"    pm_open={pm_open}  mkt_open={mkt_open}")
        print(f"    hod={hod}  lod={lod}  close={close}  day_range={day_range}")

        if not args.dry_run:
            # Only write if currently blank (never overwrite existing data)
            def write_if_blank(col, val):
                cell = ws.cell(row=row_num, column=col)
                if cell.value is None and val is not None:
                    cell.value = val

            write_if_blank(COL_PREV_HIGH,  prev_high)
            write_if_blank(COL_PREV_LOW,   prev_low)
            write_if_blank(COL_PM_OPEN,    pm_open)
            write_if_blank(COL_MKT_OPEN,   mkt_open)
            write_if_blank(COL_HOD,        hod)
            write_if_blank(COL_LOD,        lod)
            write_if_blank(COL_CLOSE,      close)
            write_if_blank(COL_DAY_RANGE,  day_range)

        filled_count += 1

    print()
    print(f"Done. Rows processed: {filled_count} | Skipped: {skipped_count} | Errors: {error_count}")

    if not args.dry_run and filled_count > 0:
        wb.save(args.file)
        print(f"Saved: {args.file}")
    elif args.dry_run:
        print("(Dry run — nothing saved)")


if __name__ == "__main__":
    main()
