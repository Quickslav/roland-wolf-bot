#!/usr/bin/env python3
"""
fill_float_shortinterest.py

Fills Short Interest + Shares Outstanding (and a short-%-of-shares-outstanding
ratio) into the EOD Winners/Losers tabs of the Roland Wolf tracker, using two
FREE, authoritative sources with strict POINT-IN-TIME matching:

  * Shares outstanding  -> SEC EDGAR XBRL (dei:EntityCommonStockSharesOutstanding)
  * Short shares        -> Nasdaq public short-interest API (bimonthly, rolling
                           ~12 months, no login; covers exchange-listed names)

DESIGN RULES (match Benjamin's standing conventions):
  * BLANK, NOT GUESSED. If a value can't be resolved for a (ticker, date) pair,
    the cell is left empty. Never 0, never carried-forward, never interpolated.
  * POINT-IN-TIME. For a trade on date D we use the most recent value that was
    already KNOWN as of D (SEC: filing filed<=D ; Nasdaq: settlement<=D), so no
    look-ahead bias creeps into the tracker.
  * MASTER IS UNTOUCHED. The script writes to a timestamped COPY, never the
    source file. Run with --dry-run first to see exactly what it would write.

IMPORTANT LIMITATION (read this):
  EDGAR gives shares OUTSTANDING, not true free float. The "Short % of Shares
  Out" column therefore UNDERSTATES true short-float for names with heavy
  insider/restricted holdings. Raw components are stored so you can recompute
  against a real float source later if you get one.

  Nasdaq short interest covers a ROLLING ~12 months, so this works for Q1-Q2
  2026 (within the window) but cannot backfill dates older than ~12 months.

ONE-TIME SETUP:
  1. SEC only requires a descriptive User-Agent (set SEC_USER_AGENT below).
  2. Nasdaq short interest needs NO account/key. (If it ever starts returning
     403/empty, run --diag to see the raw response.)
  3. pip install openpyxl requests

USAGE:
  python fill_float_shortinterest.py --diag            # check the data sources
  python fill_float_shortinterest.py --master "merged_trading_tracker.xlsx" --dry-run --limit 20
  python fill_float_shortinterest.py --master "merged_trading_tracker.xlsx"
"""

import argparse
import datetime as dt
import json
import re
import shutil
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# ----------------------------------------------------------------------------
# CONFIG  --  adjust these to match your master once its layout is confirmed
# ----------------------------------------------------------------------------

# Tabs to fill. Exact sheet names from the master.
TARGET_TABS = [
    "EOD Biggest Winners Q1", "EOD Biggest Losers Q1",
    "EOD Biggest Winners Q2", "EOD Biggest Losers Q2",
]

# Header text the script will look for (case-insensitive, first match wins).
TICKER_HEADER_CANDIDATES = ["ticker", "symbol", "stock"]
DATE_HEADER_CANDIDATES   = ["date", "trade date", "eod date", "day"]

# Output columns. The three below are inserted TOGETHER, immediately to the
# right of the column named in NEW_COLS_AFTER (so they sit next to the existing
# Float / Short Float columns rather than at the far end). If that anchor column
# isn't found, they're appended at the end instead. Re-running is safe: if the
# columns already exist they're reused in place.
COL_SHORT_SHARES = "Short Shares"
COL_SHARES_OUT   = "Shares Outstanding (SEC)"
COL_SHORT_PCT    = "Short % of Shares Out"
NEW_COLS_AFTER   = "short float"       # lower-case header text of the anchor col

# ETF flagging. Any row whose Industry/Company/News text matches is highlighted
# YELLOW so ETFs (which have no float/short-interest concept and aren't traded)
# are easy to spot and skip. Rows are flagged, never deleted.
INDUSTRY_HEADER_CANDIDATES = ["industry"]
COMPANY_HEADER_CANDIDATES  = ["company", "name", "security"]
NEWS_HEADER_CANDIDATES     = ["news title", "news", "headline"]

# Conservative markers: the exact Finviz label plus distinctive leveraged/inverse
# ETF issuer + structure terms that effectively never appear in operating-company
# names. Bare words like "bull"/"bear"/"ultra" are deliberately excluded to avoid
# false positives on real stocks.
ETF_MARKERS = [
    "exchange traded fund", " etf", " etn", "direxion", "proshares", "defiance",
    "tradr", "leverage shares", "t-rex", "graniteshares", "granite shares",
    "microsectors", "volatility shares", "teucrium", "2x long", "2x short",
    "2x daily", "3x etf", "bull 3x", "bear 3x", "-3x ", "daily target",
    "ultrapro", "proshares ultra",
]

ETF_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# SEC requires a real contact string in the User-Agent or it returns 403.
SEC_USER_AGENT = "Roland Wolf Tracker research (your_email@example.com)"

CACHE_DIR = Path(".fill_cache")
CACHE_DIR.mkdir(exist_ok=True)

SEC_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"
SEC_CONCEPT_URL = ("https://data.sec.gov/api/xbrl/companyconcept/"
                   "CIK{cik10}/dei/EntityCommonStockSharesOutstanding.json")
NASDAQ_SI_URL   = "https://api.nasdaq.com/api/quote/{symbol}/short-interest"
# Consolidated symbol directory with an ETF (Y/N) flag for every US-listed
# ticker (Nasdaq + NYSE/American/Cboe). Used to flag ETFs by ticker alone,
# so rows with no Company/Industry text still get caught.
NASDAQ_TRADED_URL = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqtraded.txt"

# Browser-like headers: api.nasdaq.com rejects requests without them.
NASDAQ_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/122.0.0.0 Safari/537.36"),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Origin": "https://www.nasdaq.com",
    "Referer": "https://www.nasdaq.com/",
}

# ----------------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------------

def norm_symbol(s):
    """Normalize a symbol for matching (uppercase, alphanumeric only)."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s else ""

def to_date(v):
    """Coerce a cell value (datetime, date, or string) to a date; None if impossible."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _num(s):
    """'1,234,567' -> 1234567 (int); '2.41' -> 2.41; blanks/dashes -> None."""
    if s is None:
        return None
    t = str(s).strip().replace(",", "").replace("$", "")
    if t in ("", "-", "--", "N/A", "NA"):
        return None
    try:
        f = float(t)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None

# ----------------------------------------------------------------------------
# SEC EDGAR : point-in-time shares outstanding
# ----------------------------------------------------------------------------

class Edgar:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers.update({"User-Agent": SEC_USER_AGENT,
                               "Accept-Encoding": "gzip, deflate"})
        self._ticker_map = None
        self._concept = {}                 # cik10 -> list of (filed, end, val)

    def _get(self, url):
        for attempt in range(4):
            try:
                r = self.s.get(url, timeout=(10, 20))
            except requests.RequestException:
                time.sleep(0.5 * (attempt + 1))   # transient network error, retry
                continue
            if r.status_code == 200:
                return r
            if r.status_code == 404:
                return None
            time.sleep(0.5 * (attempt + 1))   # rate-limit / transient backoff
        return None

    def ticker_to_cik(self, ticker):
        if self._ticker_map is None:
            cache = CACHE_DIR / "sec_tickers.json"
            if cache.exists():
                data = json.loads(cache.read_text())
            else:
                r = self._get(SEC_TICKERS_URL)
                data = r.json() if r else {}
                cache.write_text(json.dumps(data))
            self._ticker_map = {}
            for row in data.values():
                self._ticker_map[norm_symbol(row["ticker"])] = int(row["cik_str"])
        return self._ticker_map.get(norm_symbol(ticker))

    def _series(self, cik10):
        """All (filed, end, val) shares-outstanding points for a CIK, cached."""
        if cik10 in self._concept:
            return self._concept[cik10]
        cache = CACHE_DIR / f"so_{cik10}.json"
        if cache.exists():
            pts = json.loads(cache.read_text())
        else:
            r = self._get(SEC_CONCEPT_URL.format(cik10=cik10))
            pts = []
            if r:
                units = r.json().get("units", {}).get("shares", [])
                # Multiple share classes can report on the same filing; sum per
                # (end, filed) so we get TOTAL shares out, and record it.
                agg = {}
                for u in units:
                    if u.get("filed") and u.get("val") is not None:
                        key = (u.get("end"), u["filed"])
                        agg[key] = agg.get(key, 0) + int(u["val"])
                for (end, filed), val in agg.items():
                    pts.append({"end": end, "filed": filed, "val": val})
            cache.write_text(json.dumps(pts))
        self._concept[cik10] = pts
        return pts

    def shares_outstanding_asof(self, ticker, trade_date):
        """Most recent shares-outstanding value FILED on or before trade_date."""
        cik = self.ticker_to_cik(ticker)
        if cik is None:
            return None
        cik10 = str(cik).zfill(10)
        pts = self._series(cik10)
        eligible = [p for p in pts
                    if dt.date.fromisoformat(p["filed"]) <= trade_date]
        if not eligible:
            return None
        # latest by filing date, tie-break on period-end date
        best = max(eligible, key=lambda p: (p["filed"], p["end"] or ""))
        return best["val"]

# ----------------------------------------------------------------------------
# Nasdaq : point-in-time short interest (public API, bimonthly, rolling ~12 mo)
# ----------------------------------------------------------------------------

class NasdaqShortInterest:
    def __init__(self):
        self.s = requests.Session()
        self.s.headers.update(NASDAQ_HEADERS)
        self._by_symbol = {}               # norm symbol -> [(date, short_shares), ...]

    def _parse_rows(self, payload):
        """Pull [(date, short_shares)] out of a Nasdaq short-interest response.
        Defensive about key names / nesting; returns list sorted ascending."""
        data = (payload or {}).get("data") or {}
        table = data.get("shortInterestTable") or data.get("shortInterest") or {}
        rows = table.get("rows") if isinstance(table, dict) else None
        if not rows:
            return []
        out = []
        for row in rows:
            d = to_date(row.get("settlementDate") or row.get("date"))
            shares = _num(row.get("interest") or row.get("shortInterest")
                          or row.get("value"))
            if d and shares is not None:
                out.append((d, int(shares)))
        out.sort(key=lambda x: x[0])
        return out

    def _fetch_symbol(self, ticker):
        """Fetch + cache the short-interest history for one symbol."""
        sym = norm_symbol(ticker)
        if sym in self._by_symbol:
            return self._by_symbol[sym]
        cache = CACHE_DIR / f"nsi_{sym}.json"
        if cache.exists():
            hist = [(dt.date.fromisoformat(d), s)
                    for d, s in json.loads(cache.read_text())]
            self._by_symbol[sym] = hist
            return hist
        hist = []
        for asset in ("stocks", "etf"):        # try stock first, then ETF
            try:
                r = self.s.get(NASDAQ_SI_URL.format(symbol=sym),
                               params={"assetclass": asset}, timeout=(10, 20))
            except requests.RequestException:
                continue
            if r.status_code != 200 or not r.text.strip():
                continue
            try:
                payload = r.json()
            except ValueError:
                continue
            hist = self._parse_rows(payload)
            if hist:
                break
        self._by_symbol[sym] = hist
        cache.write_text(json.dumps([(d.isoformat(), s) for d, s in hist]))
        time.sleep(0.2)                        # be gentle on the public endpoint
        return hist

    def short_shares_asof(self, ticker, trade_date):
        """Short shares from the most recent settlement <= trade_date.
        Returns (short_shares, settlement_date) or (None, None)."""
        hist = self._fetch_symbol(ticker)
        eligible = [(d, s) for d, s in hist if d <= trade_date]
        if not eligible:
            return None, None
        d, s = max(eligible, key=lambda x: x[0])
        return s, d

    def diag_probe(self, symbol="AAPL"):
        """Fetch one symbol and report exactly what Nasdaq returns."""
        sym = norm_symbol(symbol)
        print(f"Nasdaq short-interest diagnostic for {sym}:")
        for asset in ("stocks", "etf"):
            try:
                r = self.s.get(NASDAQ_SI_URL.format(symbol=sym),
                               params={"assetclass": asset}, timeout=(10, 20))
            except requests.RequestException as e:
                print(f"  assetclass={asset}: request error -> {e}")
                continue
            print(f"  assetclass={asset}: status={r.status_code} "
                  f"len={len(r.text)}")
            if r.status_code != 200 or not r.text.strip():
                print(f"    body[:200]={r.text[:200]!r}")
                continue
            try:
                payload = r.json()
            except ValueError:
                print(f"    non-JSON body[:200]={r.text[:200]!r}")
                continue
            hist = self._parse_rows(payload)
            print(f"    parsed rows: {len(hist)}")
            if hist:
                print(f"    date range: {hist[0][0]} .. {hist[-1][0]}")
                print(f"    most recent 3: {hist[-3:]}")
                return
            else:
                data = (payload or {}).get("data")
                print(f"    data is null? {data is None}; "
                      f"message={payload.get('message')!r}")
        print("  (no short-interest rows parsed — see raw above)")


# ----------------------------------------------------------------------------
# ETF directory : ticker -> is-ETF, from Nasdaq's consolidated symbol directory
# ----------------------------------------------------------------------------

class EtfDirectory:
    def __init__(self):
        self._map = None                   # norm symbol -> True/False
        self.s = requests.Session()
        self.s.headers.update({"User-Agent": NASDAQ_HEADERS["User-Agent"]})

    def _load(self):
        if self._map is not None:
            return
        cache = CACHE_DIR / "etf_directory.json"
        if cache.exists():
            self._map = json.loads(cache.read_text())
            return
        self._map = {}
        try:
            r = self.s.get(NASDAQ_TRADED_URL, timeout=(10, 30))
            r.raise_for_status()
            for line in r.text.splitlines():
                parts = line.split("|")
                # fields: Traded|Symbol|Name|Exchange|MktCat|ETF|... (skip header/footer)
                if len(parts) < 7 or parts[1] in ("Symbol", ""):
                    continue
                if parts[0] == "File Creation Time":
                    continue
                self._map[norm_symbol(parts[1])] = (parts[5].strip().upper() == "Y")
            cache.write_text(json.dumps(self._map))
        except requests.RequestException as e:
            print(f"    [etf-dir] download failed ({e}); "
                  f"falling back to text-match only.")

    def is_etf(self, ticker):
        self._load()
        return bool(self._map.get(norm_symbol(ticker), False))


# ----------------------------------------------------------------------------
# spreadsheet driver
# ----------------------------------------------------------------------------

def find_header_row_and_cols(ws):
    """Scan the first 10 rows for a header row containing a ticker + date column."""
    for r in range(1, 11):
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                headers[str(v).strip().lower()] = c
        tcol = next((headers[h] for cand in TICKER_HEADER_CANDIDATES
                     for h in headers if h == cand), None)
        dcol = next((headers[h] for cand in DATE_HEADER_CANDIDATES
                     for h in headers if h == cand), None)
        if tcol and dcol:
            return r, tcol, dcol, headers
    return None, None, None, {}

def col_by_candidates(headers, candidates):
    """Return the column index for the first matching header name, or None."""
    for cand in candidates:
        if cand in headers:
            return headers[cand]
    return None

def rescan_headers(ws, header_row):
    """Map lower-cased header text -> column index for the given header row."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            headers[str(v).strip().lower()] = c
    return headers

def place_new_cols(ws, header_row, headers):
    """Ensure the three output columns exist, grouped just after the anchor
    column (NEW_COLS_AFTER). Inserts them if absent; reuses them if present.
    Returns (c_short_shares, c_shares_out, c_short_pct) and updated headers."""
    titles = [COL_SHORT_SHARES, COL_SHARES_OUT, COL_SHORT_PCT]
    keys = [t.strip().lower() for t in titles]

    if all(k in headers for k in keys):                 # already placed -> reuse
        return headers[keys[0]], headers[keys[1]], headers[keys[2]], headers

    anchor = headers.get(NEW_COLS_AFTER)
    insert_at = (anchor + 1) if anchor else (ws.max_column + 1)
    if anchor:
        ws.insert_cols(insert_at, amount=3)             # shift later cols right
    for i, title in enumerate(titles):
        ws.cell(row=header_row, column=insert_at + i, value=title)

    headers = rescan_headers(ws, header_row)            # positions changed
    return insert_at, insert_at + 1, insert_at + 2, headers

def is_etf_row(industry, company, news):
    """True if the row's text markers identify it as an ETF/ETN."""
    text = " ".join(str(x or "") for x in (industry, company, news)).lower()
    return any(m in text for m in ETF_MARKERS)

def run(master_path, dry_run, limit):
    master = Path(master_path)
    if not master.exists():
        sys.exit(f"ERROR: master not found: {master}")

    edgar, sint, etfdir = Edgar(), NasdaqShortInterest(), EtfDirectory()

    if dry_run:
        out_path = None
        wb = load_workbook(master)         # read structure; won't be saved
    else:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = master.with_name(f"{master.stem}__floatSI_{stamp}.xlsx")
        shutil.copy2(master, out_path)     # work on a copy; master is never touched
        wb = load_workbook(out_path)

    stats = {"filled_so": 0, "filled_si": 0, "blank_so": 0, "blank_si": 0,
             "rows": 0, "etf": 0}
    preview = []

    for tab in TARGET_TABS:
        if tab not in wb.sheetnames:
            print(f"  [skip] tab not found: {tab!r}")
            continue
        ws = wb[tab]
        hrow, tcol, dcol, headers = find_header_row_and_cols(ws)
        if not hrow:
            print(f"  [skip] no ticker+date header on {tab!r}")
            continue

        # columns used only for ETF detection (captured before any insertion)
        industry_col = col_by_candidates(headers, INDUSTRY_HEADER_CANDIDATES)
        company_col  = col_by_candidates(headers, COMPANY_HEADER_CANDIDATES)
        news_col     = col_by_candidates(headers, NEWS_HEADER_CANDIDATES)

        # place the three output columns grouped after Short Float
        c_ss, c_so, c_pct, headers = place_new_cols(ws, hrow, headers)

        # re-resolve everything from the (possibly shifted) header row
        tcol = col_by_candidates(headers, TICKER_HEADER_CANDIDATES) or tcol
        dcol = col_by_candidates(headers, DATE_HEADER_CANDIDATES) or dcol
        industry_col = col_by_candidates(headers, INDUSTRY_HEADER_CANDIDATES)
        company_col  = col_by_candidates(headers, COMPANY_HEADER_CANDIDATES)
        news_col     = col_by_candidates(headers, NEWS_HEADER_CANDIDATES)
        last_col = ws.max_column

        data_rows = ws.max_row - hrow
        print(f"  {tab}: ~{data_rows} rows ...", flush=True)

        for r in range(hrow + 1, ws.max_row + 1):
            ticker = ws.cell(row=r, column=tcol).value
            tdate = to_date(ws.cell(row=r, column=dcol).value)
            if not ticker or not tdate:
                continue
            if limit and stats["rows"] >= limit:
                break
            stats["rows"] += 1
            if stats["rows"] % 25 == 0:
                print(f"    ...{stats['rows']} rows done "
                      f"(so {stats['filled_so']} / si {stats['filled_si']}) "
                      f"last={ticker}", flush=True)

            # ETF check -> highlight the whole row yellow, keep processing
            industry = ws.cell(row=r, column=industry_col).value if industry_col else None
            company  = ws.cell(row=r, column=company_col).value if company_col else None
            news     = ws.cell(row=r, column=news_col).value if news_col else None
            is_etf = is_etf_row(industry, company, news) or etfdir.is_etf(ticker)
            if is_etf:
                stats["etf"] += 1
                if not dry_run:
                    for c in range(1, last_col + 1):
                        ws.cell(row=r, column=c).fill = ETF_FILL

            so = edgar.shares_outstanding_asof(ticker, tdate)
            ss, settle = sint.short_shares_asof(ticker, tdate)

            # guard bad denominators: a zero/negative shares-out is not usable,
            # so blank it (prevents #DIV/0!). BLANK, NOT GUESSED throughout.
            if so is not None and so <= 0:
                so = None

            if not dry_run:
                ws.cell(row=r, column=c_ss).value = ss if ss is not None else None
                ws.cell(row=r, column=c_so).value = so if so is not None else None
                pct_cell = ws.cell(row=r, column=c_pct)
                if ss is not None and so is not None:
                    a = f"{get_column_letter(c_ss)}{r}"
                    b = f"{get_column_letter(c_so)}{r}"
                    # formula -> stays dynamic, blank if either side missing
                    pct_cell.value = f'=IF(OR({a}="",{b}=""),"",{a}/{b})'
                    pct_cell.number_format = "0.00%"
                else:
                    pct_cell.value = None
                    pct_cell.number_format = "0.00%"

            stats["filled_so" if so is not None else "blank_so"] += 1
            stats["filled_si" if ss is not None else "blank_si"] += 1
            if len(preview) < 15:
                pct = f"{ss/so:.2%}" if (ss and so) else "—"
                preview.append((tab, str(ticker), tdate.isoformat(),
                                so if so is not None else "—",
                                ss if ss is not None else "—", pct,
                                settle.isoformat() if settle else "—"))
        if limit and stats["rows"] >= limit:
            break

    print("\n  sample of resolved values (tab | ticker | date | sharesOut | shortShares | short% | SI settle):")
    for row in preview:
        print("   ", " | ".join(str(x) for x in row))
    print(f"\n  rows processed        : {stats['rows']}")
    print(f"  shares-out filled/blank: {stats['filled_so']} / {stats['blank_so']}")
    print(f"  short-int filled/blank : {stats['filled_si']} / {stats['blank_si']}")
    print(f"  ETF rows highlighted   : {stats['etf']}")

    if dry_run:
        print("\n  DRY RUN — nothing written. Re-run without --dry-run to write a copy.")
    else:
        wb.save(out_path)
        print(f"\n  WROTE COPY: {out_path}")
        print("  (Master untouched. Open in Excel to let the % formulas calculate,")
        print("   or run it through LibreOffice recalc if you want cached values.)")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--master", help="path to the tracker .xlsx")
    ap.add_argument("--dry-run", action="store_true",
                    help="resolve values and report, but write nothing")
    ap.add_argument("--limit", type=int, default=0,
                    help="cap rows processed (for a quick test)")
    ap.add_argument("--diag", action="store_true",
                    help="probe the data sources and exit")
    ap.add_argument("--diag-symbol", default="AAPL",
                    help="ticker to test for the Nasdaq probe (default AAPL)")
    args = ap.parse_args()
    if args.diag:
        print(f"SEC user-agent: {SEC_USER_AGENT!r}")
        if "your_email@example.com" in SEC_USER_AGENT:
            print("  WARNING: set a real email in SEC_USER_AGENT or EDGAR may 403.\n")
        # quick SEC check: resolve a known ticker to a CIK
        try:
            cik = Edgar().ticker_to_cik(args.diag_symbol)
            print(f"SEC ticker->CIK for {args.diag_symbol}: {cik}\n")
        except Exception as e:
            print(f"SEC check error -> {e}\n")
        NasdaqShortInterest().diag_probe(args.diag_symbol)
        # ETF directory check
        print()
        ed = EtfDirectory()
        ed._load()
        n = len(ed._map or {})
        print(f"ETF directory: {n} symbols loaded")
        if n:
            for t in ("AAPL", "SOXS", "SOXL", "MSTU", "IONZ", args.diag_symbol):
                print(f"  {t}: is_etf={ed.is_etf(t)}")
        return
    if not args.master:
        sys.exit("ERROR: --master is required (or use --diag).")
    run(args.master, args.dry_run, args.limit)


if __name__ == "__main__":
    main()
