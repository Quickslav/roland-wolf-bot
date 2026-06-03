"""
add_close_prices.py
--------------------
Fetches market close prices from Alpaca for every ticker/date in the
Pre-Trade Data tab, then inserts a new "Close Price ($)" column after
column P (Day Range) in the spreadsheet.

No other data is modified.

Usage:
    pip install requests openpyxl
    python add_close_prices.py

Edit the three config lines below if needed.
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from collections import defaultdict

# ── Config ──────────────────────────────────────────────────────────────
INPUT_FILE  = "merged_trading_tracker__1_.xlsx"
OUTPUT_FILE = "merged_trading_tracker__1_.xlsx"   # overwrite in place
API_KEY     = "PKIYGXQT3DGX7B6BDIZFZ6VQWU"
API_SECRET  = "Ekaz5bQHUbbFUQvmidbBSU89wHMtgigik3TsyFD15NA3"
SHEET_NAME  = "Pre-Trade Data"
# ────────────────────────────────────────────────────────────────────────

BASE_URL = "https://data.alpaca.markets/v2/stocks/bars"
HEADERS  = {
    "APCA-API-KEY-ID":     API_KEY,
    "APCA-API-SECRET-KEY": API_SECRET,
}

def parse_date(raw):
    if isinstance(raw, datetime):
        return raw.date()
    s = str(raw).strip()
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None

def fetch_closes(tickers, start, end):
    """Returns {(ticker, date): close_price}"""
    result = {}
    # Alpaca accepts up to ~1000 symbols per request; batch in 200s
    ticker_list = sorted(tickers)
    batch_size  = 200
    start_str   = start.strftime("%Y-%m-%d")
    end_str     = (end + timedelta(days=1)).strftime("%Y-%m-%d")

    for i in range(0, len(ticker_list), batch_size):
        batch = ticker_list[i : i + batch_size]
        params = {
            "symbols":   ",".join(batch),
            "timeframe": "1Day",
            "start":     start_str,
            "end":       end_str,
            "limit":     10000,
            "feed":      "iex",        # free-tier feed
            "adjustment":"raw",
        }
        resp = requests.get(BASE_URL, headers=HEADERS, params=params, timeout=30)
        resp.raise_for_status()
        data = resp.json().get("bars", {})
        for ticker, bars in data.items():
            for bar in bars:
                bar_date = datetime.fromisoformat(bar["t"].replace("Z","")).date()
                result[(ticker, bar_date)] = round(bar["c"], 4)
        # handle pagination
        next_token = resp.json().get("next_page_token")
        while next_token:
            params["page_token"] = next_token
            resp = requests.get(BASE_URL, headers=HEADERS, params=params, timeout=30)
            resp.raise_for_status()
            js   = resp.json()
            for ticker, bars in js.get("bars", {}).items():
                for bar in bars:
                    bar_date = datetime.fromisoformat(bar["t"].replace("Z","")).date()
                    result[(ticker, bar_date)] = round(bar["c"], 4)
            next_token = js.get("next_page_token")

    return result

def main():
    print(f"Loading {INPUT_FILE} ...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb[SHEET_NAME]

    # ── Read all data rows to find ticker/date pairs ──────────────────
    all_rows = list(ws.iter_rows(min_row=3))   # row 1=title, row 2=header
    entries  = []   # (excel_row_number, date, ticker)
    tickers  = set()
    dates    = []

    for row_cells in all_rows:
        raw_date = row_cells[0].value
        ticker   = row_cells[1].value
        if not raw_date or not ticker:
            continue
        date = parse_date(raw_date)
        if date is None:
            continue
        excel_row = row_cells[0].row
        entries.append((excel_row, date, ticker))
        tickers.add(ticker)
        dates.append(date)

    if not entries:
        print("No data rows found — check sheet name.")
        return

    start_date = min(dates)
    end_date   = max(dates)
    print(f"Found {len(entries)} rows | {len(tickers)} unique tickers | {start_date} → {end_date}")

    # ── Fetch prices ──────────────────────────────────────────────────
    print("Fetching close prices from Alpaca ...")
    closes = fetch_closes(tickers, start_date, end_date)
    print(f"Received {len(closes)} price points")

    # ── Insert new column Q (index 17, after column P = index 16) ────
    # Column P is index 17 (1-based), so we insert at column 17 to push
    # everything from column Q onward right by one.
    INSERT_COL = 17   # inserts BEFORE current col 17, making it the new col 17

    ws.insert_cols(INSERT_COL)

    # ── Write header in row 2 ─────────────────────────────────────────
    header_cell = ws.cell(row=2, column=INSERT_COL)
    header_cell.value = "Close Price ($)"
    # Match the style of neighbouring header cells
    header_cell.font      = Font(bold=True)
    header_cell.alignment = Alignment(horizontal="center")

    # ── Write close prices ────────────────────────────────────────────
    found = 0
    missing = []
    for (excel_row, date, ticker) in entries:
        price = closes.get((ticker, date))
        cell  = ws.cell(row=excel_row, column=INSERT_COL)
        if price is not None:
            cell.value        = price
            cell.number_format = '#,##0.00'
            found += 1
        else:
            cell.value = "N/A"
            missing.append(f"  {ticker} {date}")

    print(f"\nPrices written : {found}/{len(entries)}")
    if missing:
        print(f"Not found ({len(missing)}):")
        for m in missing:
            print(m)

    # ── Save ──────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\nSaved → {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
